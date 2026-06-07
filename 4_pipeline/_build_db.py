#!/usr/bin/env python3
"""Build summit_targets.db (SQLite) from the master workbook.

- One `companies` table; column names preserved (quoted identifiers).
- Values inserted as-is for fidelity (datetimes -> ISO strings).
- Indexes on hierarchy + screening keys for fast slicing.
"""
import openpyxl, sqlite3, os, time, datetime

MASTER = "1_source_data/Data Axle Master - All Transport Companies.xlsx"
DB = "2_database/summit_targets.db"
BATCH = 5000
INDEX_COLS = ["IUSA Number", "Parent IUSA Number", "Subsidiary IUSA Number",
              "Parent Company Name", "Company Name", "Location Type",
              "Firm or Individual", "Primary NAICS", "Primary SIC Code",
              "State", "Source Batch"]

def cell(v):
    if v is None or isinstance(v, (str, int, float, bytes)):
        return v
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v)

def main():
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if os.path.exists(DB):
        os.remove(DB)
    wb = openpyxl.load_workbook(MASTER, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rit = ws.iter_rows(values_only=True)
    header = list(next(rit))
    ncols = len(header)
    print(f"[{time.strftime('%H:%M:%S')}] {ncols} columns, sheet '{ws.title}'", flush=True)

    con = sqlite3.connect(DB)
    cur = con.cursor()
    cur.execute("PRAGMA journal_mode=OFF")
    cur.execute("PRAGMA synchronous=OFF")
    coldefs = ", ".join(f'"{h}"' for h in header)
    cur.execute(f"CREATE TABLE companies ({coldefs})")
    ins = f"INSERT INTO companies VALUES ({','.join('?'*ncols)})"

    n = 0
    batch = []
    for row in rit:
        row = list(row)
        if len(row) < ncols:
            row += [None] * (ncols - len(row))
        batch.append([cell(v) for v in row])
        if len(batch) >= BATCH:
            cur.executemany(ins, batch); n += len(batch); batch = []
            if n % 50000 == 0:
                print(f"[{time.strftime('%H:%M:%S')}] inserted {n:,}", flush=True)
    if batch:
        cur.executemany(ins, batch); n += len(batch)
    con.commit()
    wb.close()
    print(f"[{time.strftime('%H:%M:%S')}] inserted total {n:,}; building indexes", flush=True)

    for col in INDEX_COLS:
        safe = col.replace(" ", "_").replace("/", "_")
        cur.execute(f'CREATE INDEX "idx_{safe}" ON companies ("{col}")')
    con.commit()
    cur.execute("PRAGMA optimize")
    con.close()
    print(f"[{time.strftime('%H:%M:%S')}] DONE", flush=True)
    print(f"  Rows  : {n:,}", flush=True)
    print(f"  Cols  : {ncols}", flush=True)
    print(f"  DB    : {DB} ({os.path.getsize(DB)/1e6:.1f} MB)", flush=True)

if __name__ == "__main__":
    main()
